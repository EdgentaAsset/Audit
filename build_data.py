#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build_data.py  —  Tukar Data.xlsx -> app-data.js untuk App Audit Aset HoSZA.

Guna:
    python build_data.py

Output:
    app-data.js   (window.APP_DATA = {...})

Jalankan semula fail ini setiap kali Data.xlsx dikemas kini.
"""
import json
import datetime
import os
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(HERE, "Data.xlsx")
OUT = os.path.join(HERE, "app-data.js")

# Susunan medan ringkas untuk setiap aset (array-of-arrays untuk jimat saiz).
# Index ini mesti sepadan dengan FIELDS dalam index.html.
#  0 ASSET NO        1 NO. UNIZA      2 TYPE DESCRIPTION  3 TYPE CODE
#  4 USER DEPARTMENT 5 DEPARTMENT NAME 6 LOCATION NO.     7 LOCATION NAME
#  8 WORKGROUP       9 MAKE          10 BRAND            11 MODEL
# 12 SERIAL NO.     13 MANUFACTURER  14 GAMBAR 1 (Aset)  15 GAMBAR 2 (Plate)

ASSET_COL = {
    "asset": 0, "g1": 1, "g2": 2, "uniza": 3, "typedesc": 4, "typecode": 5,
    "userdept": 6, "deptname": 7, "locno": 8, "locname": 9, "workgroup": 10,
    "make": 11, "brand": 12, "model": 13, "serial": 14, "manuf": 15,
}


def clean(v):
    if v is None:
        return ""
    s = str(v).strip()
    if s.lower() in ("n/a", "na", "not available", "-", "none"):
        return s  # kekalkan apa adanya supaya admin nampak nilai asal
    return s


def load_assets(wb):
    ws = wb["AssetDetails"]
    out = []
    for r in ws.iter_rows(min_row=4, values_only=True):
        akey = r[ASSET_COL["asset"]]
        if not akey or not str(akey).strip():
            continue
        out.append([
            clean(r[ASSET_COL["asset"]]),
            clean(r[ASSET_COL["uniza"]]),
            clean(r[ASSET_COL["typedesc"]]),
            clean(r[ASSET_COL["typecode"]]),
            clean(r[ASSET_COL["userdept"]]),
            clean(r[ASSET_COL["deptname"]]),
            clean(r[ASSET_COL["locno"]]),
            clean(r[ASSET_COL["locname"]]),
            clean(r[ASSET_COL["workgroup"]]),
            clean(r[ASSET_COL["make"]]),
            clean(r[ASSET_COL["brand"]]),
            clean(r[ASSET_COL["model"]]),
            clean(r[ASSET_COL["serial"]]),
            clean(r[ASSET_COL["manuf"]]),
            clean(r[ASSET_COL["g1"]]),
            clean(r[ASSET_COL["g2"]]),
        ])
    return out


def fmt_date(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    return s


def load_ppm(wb):
    ws = wb["PPM Schedule"]
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    # Baris 1 (index 0) = label julat minggu utk lajur 8..59
    week_labels = [clean(rows[0][i]) for i in range(8, 60)]
    ppm = {}  # asset -> {"s": 0/1, "t": [ [disc,desc,task,typecode,cat,ppmType,ppm,[ [wkLabel,date], ... ]] ] }
    for r in rows[2:]:  # data mula baris 3 (index 2)
        akey = r[0]
        if not akey or not str(akey).strip():
            continue
        akey = str(akey).strip()
        dates = []
        for i in range(8, 60):
            cell = r[i]
            if cell not in (None, ""):
                dates.append([week_labels[i - 8], fmt_date(cell)])
        is_sched = 1 if str(r[7]).strip().lower() == "scheduled" else 0
        task = [
            clean(r[1]),  # Discipline
            clean(r[2]),  # Asset Description
            clean(r[3]),  # Task Code
            clean(r[4]),  # Type Code
            clean(r[5]),  # Category Description
            clean(r[6]),  # PPM Type
            clean(r[7]),  # PPM (scheduled/Unscheduled)
            dates,
        ]
        rec = ppm.setdefault(akey, {"s": 0, "t": []})
        rec["t"].append(task)
        if is_sched:
            rec["s"] = 1
    return ppm


def main():
    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    assets = load_assets(wb)
    ppm = load_ppm(wb)

    payload = {
        "version": datetime.datetime.now().strftime("%Y%m%d%H%M"),
        "fields": ["asset", "uniza", "typedesc", "typecode", "userdept",
                   "deptname", "locno", "locname", "workgroup", "make",
                   "brand", "model", "serial", "manuf", "g1", "g2"],
        "assets": assets,
        "ppm": ppm,
    }

    body = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
    with open(OUT, "w", encoding="utf-8") as f:
        f.write("window.APP_DATA=")
        f.write(body)
        f.write(";")

    size = os.path.getsize(OUT)
    print("Assets        :", len(assets))
    print("PPM assets    :", len(ppm))
    print("Output        :", OUT)
    print("Size          : %.2f MB (%d bytes)" % (size / 1048576, size))


if __name__ == "__main__":
    main()
